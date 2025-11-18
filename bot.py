import telebot
from telebot import types
import openpyxl
import datetime

# ================================
#    НАЛАШТУВАННЯ
# ================================
BOT_TOKEN = "8467962731:AAEn7fHQwJ9LN-EmVWIJoITqrmE_BkT2Z_s"
ADMIN_ID = 123456789   # Сюди бот надсилає всі заявки

bot = telebot.TeleBot(BOT_TOKEN)

# ================================
#    СТВОРЕННЯ ФАЙЛУ ЗАЯВОК
# ================================
def init_excel():
    try:
        book = openpyxl.load_workbook("clients.xlsx")
    except:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet["A1"] = "Дата"
        sheet["B1"] = "Ім'я"
        sheet["C1"] = "Телефон"
        sheet["D1"] = "Послуга"
        sheet["E1"] = "Коментар"
        book.save("clients.xlsx")

init_excel()

# ================================
#    МЕНЮ
# ================================
def main_menu():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton("📌 Про компанію")
    btn2 = types.KeyboardButton("💬 Консультація")
    btn3 = types.KeyboardButton("📞 Контакти")
    btn4 = types.KeyboardButton("📝 Залишити заявку")
    return markup.add(btn1, btn2, btn3, btn4)


# ================================
#    КОМАНДА /start
# ================================
@bot.message_handler(commands=['start'])
def start(message):
    bot.send_message(
        message.chat.id,
        "Вітаю! 👋\n\nЯ чат-бот компанії. Виберіть дію нижче:",
        reply_markup=main_menu(),
    )

# ================================
#    ОБРОБКА КНОПОК МЕНЮ
# ================================
@bot.message_handler(func=lambda msg: msg.text in [
    "📌 Про компанію", "💬 Консультація",
    "📞 Контакти", "📝 Залишити заявку"
])
def menu_handler(message):

    if message.text == "📌 Про компанію":
        bot.send_message(message.chat.id,
            "Наша компанія займається професійними послугами.\n"
            "Пишіть — і ми допоможемо!"
        )

    elif message.text == "💬 Консультація":
        bot.send_message(message.chat.id,
            "Поставте будь-яке питання — менеджер відповість."
        )

    elif message.text == "📞 Контакти":
        bot.send_message(message.chat.id,
            "📞 Телефон: +38 099 123 45 67\n"
            "🌐 Сайт: https://example.com\n"
            "📍 Адреса: Київ"
        )

    elif message.text == "📝 Залишити заявку":
        msg = bot.send_message(message.chat.id, "Введіть ваше ім’я:")
        bot.register_next_step_handler(msg, get_name)

# ================================
#    ФОРМА ЗАЯВКИ
# ================================
def get_name(message):
    name = message.text

    msg = bot.send_message(message.chat.id, "Ваш номер телефону:")
    bot.register_next_step_handler(msg, get_phone, name)

def get_phone(message, name):
    phone = message.text

    msg = bot.send_message(message.chat.id, "Яку послугу ви хочете?")
    bot.register_next_step_handler(msg, get_service, name, phone)

def get_service(message, name, phone):
    service = message.text

    msg = bot.send_message(message.chat.id, "Ваш коментар:")
    bot.register_next_step_handler(msg, finish_form, name, phone, service)

def finish_form(message, name, phone, service):
    comment = message.text

    # ======= запис у Excel =======
    book = openpyxl.load_workbook("clients.xlsx")
    sheet = book.active
    row = [datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
           name, phone, service, comment]
    sheet.append(row)
    book.save("clients.xlsx")

    # ======= надсилання адміну =======
    bot.send_message(
        ADMIN_ID,
        f"🔥 Нова заявка!\n\n"
        f"👤 Ім'я: {name}\n"
        f"📞 Телефон: {phone}\n"
        f"💼 Послуга: {service}\n"
        f"💬 Коментар: {comment}"
    )

    # ======= відповідь клієнту =======
    bot.send_message(
        message.chat.id,
        "Дякую! Ваша заявка успішно надіслана. Менеджер скоро зв’яжеться з вами.",
        reply_markup=main_menu()
    )


# ================================
#    ОБРОБНИК ПОМИЛОК
# ================================
@bot.message_handler(func=lambda msg: True)
def fallback(message):
    bot.send_message(
        message.chat.id,
        "Не розумію команду 😕\nВиберіть пункт із меню нижче:",
        reply_markup=main_menu()
    )

# ================================
#    ЗАПУСК
# ================================
print("Bot is running...")
bot.infinity_polling()
